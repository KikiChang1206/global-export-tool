import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from collections import OrderedDict
import math

# ── 網頁基本設定 ──────────────────────────────────────────────
st.set_page_config(page_title="✈️ 全球出口轉換工具", layout="centered")
st.markdown("""
<style>
.stApp { background: linear-gradient(135deg, #0f2027, #203a43, #2c5364); }
.main-title {
    font-size: 32px; font-weight: 900; color: #FFFFFF;
    text-align: center; letter-spacing: 2px;
    text-shadow: 0 0 20px rgba(100,220,255,0.6);
    padding: 10px 0 4px 0;
}
.sub-title {
    font-size: 13px; color: #90caf9; text-align: center;
    letter-spacing: 1px; margin-bottom: 28px;
}
.block-card {
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.15);
    border-radius: 16px;
    padding: 24px 28px 20px 28px;
    margin-bottom: 24px;
    backdrop-filter: blur(6px);
}
.card-title {
    font-size: 17px; font-weight: 700; color: #64dfdf;
    margin-bottom: 14px; letter-spacing: 1px;
}
.stFileUploader section {
    background: rgba(255,255,255,0.08) !important;
    border: 1.5px dashed rgba(100,223,223,0.5) !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploadDropzone"] p { color: #b0bec5 !important; }
div.stButton > button {
    background: linear-gradient(90deg, #00b4d8, #0077b6) !important;
    color: #FFFFFF !important; border: none !important;
    border-radius: 10px !important; height: 48px;
    font-size: 15px; font-weight: 700; width: 100%;
    letter-spacing: 1px; box-shadow: 0 4px 15px rgba(0,180,216,0.4);
}
[data-testid="stDownloadButton"] button {
    background: linear-gradient(90deg, #06d6a0, #048a81) !important;
    color: #FFFFFF !important; border: none !important;
    border-radius: 10px !important; height: 48px;
    font-size: 15px; font-weight: 700; width: 100%;
    box-shadow: 0 4px 15px rgba(6,214,160,0.4);
}
footer { visibility: hidden; }
label, .stMarkdown p { color: #cfd8dc !important; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">✈️ 全球出口轉換工具</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">SFL Logistics｜Packing List & Invoice 自動整理系統</div>', unsafe_allow_html=True)

# ── 箱子尺寸對照 ──────────────────────────────────────────────
def get_dimensions(pcs):
    try:
        val = int(float(pcs))
        if   1  <= val <= 5:  return "18*19*15"
        elif 6  <= val <= 10: return "23*14*14"
        elif 11 <= val <= 40: return "29*20*20"
        elif val >= 41:       return "42*30*25"
        return ""
    except:
        return ""

# ── 讀檔共用函式 ──────────────────────────────────────────────
def read_df(uploaded_file):
    engine = 'xlrd' if uploaded_file.name.lower().endswith('.xls') else 'openpyxl'
    xl = pd.ExcelFile(uploaded_file, engine=engine)
    sheet = 'ItemData' if 'ItemData' in xl.sheet_names else xl.sheet_names[0]
    return xl.parse(sheet, header=None, dtype=str).fillna('')

# ── 表頭擷取共用函式 ──────────────────────────────────────────
def extract_header(df):
    h = {}
    h['r1_A']  = str(df.iloc[0, 0]).strip()
    h['r2_A']  = str(df.iloc[1, 0]).strip()
    h['r3_A']  = str(df.iloc[2, 0]).strip()
    h['r3_E']  = str(df.iloc[2, 4]).strip()
    h['r4_A']  = str(df.iloc[3, 0]).strip()
    h['r5_A']  = str(df.iloc[4, 0]).strip()
    h['r5_E']  = str(df.iloc[4, 4]).strip()
    h['r6_A']  = str(df.iloc[5, 0]).strip()
    h['r7_A']  = str(df.iloc[6, 0]).strip()
    h['r7_E']  = str(df.iloc[6, 4]).strip()
    h['r8_A']  = str(df.iloc[7, 0]).strip()
    h['r8_E']  = str(df.iloc[7, 4]).strip()
    h['r9_A']  = str(df.iloc[8, 0]).strip()
    h['r9_D']  = str(df.iloc[8, 3]).strip()
    h['r9_G']  = str(df.iloc[8, 6]).strip()
    h['r10_A'] = str(df.iloc[9, 0]).strip()
    h['r10_E'] = str(df.iloc[9, 4]).strip()
    return h

# ── 表頭寫入 + 合併 + 列高 共用函式 ──────────────────────────
def write_header(ws, h, bold12, total_cols):
    last = chr(64 + total_cols)

    def sc(row, col, val, wrap=False, ha='left'):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = bold12
        cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)

    sc(1,  1, h['r1_A'],  ha='center')
    sc(2,  1, h['r2_A'],  ha='center')
    sc(3,  1, h['r3_A']);  sc(3,  4, h['r3_E'])
    sc(4,  1, h['r4_A'],  ha='center')
    sc(5,  1, h['r5_A']);  sc(5,  4, h['r5_E'])
    sc(6,  1, h['r6_A'])
    sc(7,  1, h['r7_A']);  sc(7,  4, h['r7_E'], wrap=True)
    sc(8,  1, h['r8_A'],  wrap=True); sc(8, 4, h['r8_E'])
    sc(9,  1, h['r9_A']);  sc(9,  3, h['r9_D']); sc(9, 6, h['r9_G'])
    sc(10, 1, h['r10_A']); sc(10, 4, h['r10_E'])

    for rng, ha in [
        (f'A1:{last}1','center'), (f'A2:{last}2','center'), (f'A4:{last}4','center'),
        ('A3:C3','left'),  (f'D3:{last}3','left'),
        ('A5:C5','left'),  (f'D5:{last}5','left'),
        (f'A6:{last}6','left'),
        ('A7:C7','left'),  (f'D7:{last}7','left'),
        ('A8:C8','left'),  (f'D8:{last}8','left'),
        ('A9:B9','left'),  ('C9:E9','left'),
        ('A10:C10','left'),(f'D10:{last}10','left'),
    ]:
        ws.merge_cells(rng)
        tl = rng.split(':')[0]
        ew = ws[tl].alignment.wrap_text
        ws[tl].alignment = Alignment(horizontal=ha, vertical='center', wrap_text=ew)

    # 列高：1~6=15.5，7=77.7，8~12=15.5
    for r in range(1, 13):
        if r == 7:
            ws.row_dimensions[r].height = 77.7
        else:
            ws.row_dimensions[r].height = 15.5

    # 字型補套
    for r in range(1, 13):
        for c in range(1, total_cols + 1):
            ws.cell(r, c).font = bold12

# ── Packing 轉換函式 ──────────────────────────────────────────
def convert_packing(uploaded_file):
    df = read_df(uploaded_file)
    h  = extract_header(df)

    items = []
    for idx in range(12, len(df)):
        col0 = str(df.iloc[idx, 0]).strip()
        if "總箱數" in col0 or "TOTAL" in col0.upper():
            break
        desc_zh = str(df.iloc[idx, 2]).strip()
        if "SHIPFEE" in desc_zh.upper():
            continue
        items.append(df.iloc[idx].tolist())

    groups, current = [], []
    for row in items:
        sku = str(row[0]).strip()
        if sku:
            if current: groups.append(current)
            current = [row]
        else:
            current.append(row)
    if current: groups.append(current)

    final_rows = []
    total_qty = total_net = total_gross = 0.0
    sku_count = len(groups)

    for group in groups:
        group_qty = 0.0
        for r in group:
            try: group_qty += float(r[3])
            except: pass
        total_qty   += group_qty
        gross_w      = round(group_qty * 0.1, 2)
        net_w        = max(0.0, round(gross_w - 0.05, 2))
        total_gross += gross_w
        total_net   += net_w
        meas         = get_dimensions(group_qty)
        for i, r in enumerate(group):
            desc = str(r[2])
            for kw in ["【新品】", "★", "【歡樂智多星推薦】"]:
                desc = desc.replace(kw, "")
            desc = desc.strip()
            final_rows.append([
                r[0] if i == 0 else "", desc, r[3],
                net_w   if i == 0 else "",
                gross_w if i == 0 else "",
                meas    if i == 0 else "",
            ])

    wb = Workbook(); ws = wb.active; ws.title = "Packing"
    bold12      = Font(name='Calibri', size=12, bold=True)
    thin_s      = Side(style='thin', color='000000')
    thin_border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    for col, w in {'A':22.9,'B':37.09,'C':6.1,'D':20.5,'E':22.9,'F':21.7}.items():
        ws.column_dimensions[col].width = w

    write_header(ws, h, bold12, 6)

    for c, val in enumerate(["SKU","Description_of_Goods_(zh)","Qty","Net_Weight_(KG)","Gross_Weight_(KG)","Measurement_(cm)"], 1):
        cell = ws.cell(row=12, column=c, value=val)
        cell.font = bold12; cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for i, row_data in enumerate(final_rows):
        r = 13 + i
        # 計算各欄需要的行數（用像素估算，Calibri 12pt bold）
        def px_lines(text, col_width):
            if not text: return 1
            col_px = col_width * 7.0
            text_px = sum(17 if ord(c) > 127 else 9 for c in str(text))
            return max(1, math.ceil(text_px / col_px))
        lines = px_lines(row_data[1], 37.09)  # B: 品名決定列高
        ws.row_dimensions[r].height = max(15.5, lines * 16.5)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = bold12; cell.border = thin_border
            if c == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif c <= 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if val != "" and c in [3, 4, 5]:
                try:
                    cell.value = float(val)
                    cell.number_format = '0' if c == 3 else '0.00'
                except: pass

    total_row = 13 + len(final_rows)
    ws.row_dimensions[total_row].height = 15.5
    qty_disp = int(total_qty) if total_qty == int(total_qty) else round(total_qty, 2)
    for c, val in enumerate([f"總箱數:{sku_count}箱","", qty_disp, round(total_net,2), round(total_gross,2),""], 1):
        cell = ws.cell(row=total_row, column=c, value=val)
        cell.font = bold12; cell.border = thin_border
        cell.alignment = Alignment(horizontal='left' if c<=3 else 'center', vertical='center')

    output = BytesIO(); wb.save(output)
    return output.getvalue()

# ── Invoice 轉換函式 ──────────────────────────────────────────
def convert_invoice(uploaded_file):
    df = read_df(uploaded_file)
    h  = extract_header(df)

    raw_items = []
    for idx in range(12, len(df)):
        col0 = str(df.iloc[idx, 0]).strip()
        if col0 == '':
            amt = str(df.iloc[idx, 8]).strip()
            if amt and 'TWD' in amt.upper(): break
            continue
        desc_zh = str(df.iloc[idx, 2]).strip()
        if 'SHIPFEE' in desc_zh.upper(): continue
        for kw in ['【新品】', '★', '【歡樂智多星推薦】']:
            desc_zh = desc_zh.replace(kw, '')
        desc_zh = desc_zh.strip()
        try: qty    = float(str(df.iloc[idx, 5]).strip())
        except: qty = 0.0
        try: amount    = float(str(df.iloc[idx, 8]).strip())
        except: amount = 0.0
        origin = str(df.iloc[idx, 4]).strip()
        raw_items.append({'sku': col0, 'desc': desc_zh, 'origin': origin, 'qty': qty, 'amount': amount})

    groups = OrderedDict()
    for item in raw_items:
        sku = item['sku']
        if sku not in groups:
            groups[sku] = {'desc': item['desc'], 'origin': item['origin'], 'qty': 0.0, 'amount': 0.0}
        groups[sku]['qty']    += item['qty']
        groups[sku]['amount'] += item['amount']

    final_rows = []
    total_qty = total_amount = 0.0
    for sku, g in groups.items():
        qty          = g['qty']
        amount       = g['amount']
        unit_price   = math.floor((amount / qty) + 0.5) if qty > 0 else 0
        final_amount = qty * unit_price
        brand        = 'Shalom希樂' if '益生菌' in g['desc'] else 'Jealousness婕洛妮絲'
        total_qty    += qty
        total_amount += final_amount
        final_rows.append([sku, g['desc'], brand, g['origin'], qty, unit_price, final_amount])

    wb = Workbook(); ws = wb.active; ws.title = "Invoice"
    bold12      = Font(name='Calibri', size=12, bold=True)
    thin_s      = Side(style='thin', color='000000')
    thin_border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    for col, w in {'A':22.9,'B':35.64,'C':24.5,'D':9.7,'E':7.5,'F':14.5,'G':17.09}.items():
        ws.column_dimensions[col].width = w

    write_header(ws, h, bold12, 7)

    for c, val in enumerate(['SKU','Description_of_Goods_(zh)','Brand','Origin','Qty','Unit_Price','Amount'], 1):
        cell = ws.cell(row=12, column=c, value=val)
        cell.font = bold12; cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for i, row_data in enumerate(final_rows):
        r = 13 + i
        # 計算各欄需要的行數（用像素估算，Calibri 12pt bold）
        def px_lines(text, col_width):
            if not text: return 1
            col_px = col_width * 7.0
            text_px = sum(17 if ord(c) > 127 else 9 for c in str(text))
            return max(1, math.ceil(text_px / col_px))
        lines = px_lines(row_data[1], 35.64)  # B: 品名決定列高
        ws.row_dimensions[r].height = max(15.5, lines * 16.5)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = bold12; cell.border = thin_border
            if c == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif c <= 4:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if c in [5, 6, 7] and val != '':
                try:
                    cell.value = float(val)
                    cell.number_format = '0' if c in [5, 6] else '0.00'
                except: pass

    total_row = 13 + len(final_rows)
    ws.row_dimensions[total_row].height = 15.5
    qty_disp = int(total_qty) if total_qty == int(total_qty) else round(total_qty, 2)
    for c, val in enumerate([f'Total:{len(final_rows)}項','','','', qty_disp,'', round(total_amount,2)], 1):
        cell = ws.cell(row=total_row, column=c, value=val)
        cell.font = bold12; cell.border = thin_border
        cell.alignment = Alignment(horizontal='left' if c<=4 else 'center', vertical='center')
        if c == 7 and val != '': cell.number_format = '0.00'

    output = BytesIO(); wb.save(output)
    return output.getvalue()

# ── Streamlit UI ──────────────────────────────────────────────
st.markdown('<div class="block-card"><div class="card-title">📂 上傳檔案（可同時選取兩個）</div>', unsafe_allow_html=True)
uploaded_files = st.file_uploader(
    "系統會自動判斷 MergePackingList / MergeInvoice，一次選兩個也沒問題",
    type=['xls', 'xlsx'],
    accept_multiple_files=True
)
st.markdown('</div>', unsafe_allow_html=True)

if st.button("🚀 一鍵執行轉換", use_container_width=True):
    if not uploaded_files:
        st.warning("⚠️ 請先上傳檔案")
    else:
        packing_file = None
        invoice_file = None
        unrecognized = []

        for f in uploaded_files:
            name = f.name.upper()
            if 'MERGEPACKINGLIST' in name:
                packing_file = f
            elif 'MERGEINVOICE' in name:
                invoice_file = f
            else:
                unrecognized.append(f.name)

        if unrecognized:
            st.warning(f"⚠️ 無法判斷類型（檔名不含 MergePackingList 或 MergeInvoice）：{', '.join(unrecognized)}")

        if packing_file:
            try:
                with st.spinner("📦 Packing 處理中..."):
                    result_p = convert_packing(packing_file)
                st.session_state['packing_result'] = result_p
                st.session_state['packing_name']   = packing_file.name.replace('.xls','').replace('.xlsx','') + '_output.xlsx'
                st.success("✅ Packing 轉換完成！")
            except Exception as e:
                st.error(f"Packing 錯誤：{e}")

        if invoice_file:
            try:
                with st.spinner("🧾 Invoice 處理中..."):
                    result_i = convert_invoice(invoice_file)
                st.session_state['invoice_result'] = result_i
                st.session_state['invoice_name']   = invoice_file.name.replace('.xls','').replace('.xlsx','') + '_output.xlsx'
                st.success("✅ Invoice 轉換完成！")
            except Exception as e:
                st.error(f"Invoice 錯誤：{e}")

# 下載區：session_state 有資料就顯示，不會消失
if st.session_state.get('packing_result') or st.session_state.get('invoice_result'):
    st.markdown("---")
    st.markdown('<div class="card-title">📥 下載區</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        if st.session_state.get('packing_result'):
            st.download_button(
                label="📦 下載 Packing",
                data=st.session_state['packing_result'],
                file_name=st.session_state.get('packing_name', 'Packing_output.xlsx'),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_packing"
            )
    with col2:
        if st.session_state.get('invoice_result'):
            st.download_button(
                label="🧾 下載 Invoice",
                data=st.session_state['invoice_result'],
                file_name=st.session_state.get('invoice_name', 'Invoice_output.xlsx'),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_invoice"
            )
